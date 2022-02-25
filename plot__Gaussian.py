import matplotlib as mpl
import matplotlib.pyplot as plt
import openpyxl as pyxl
import numpy as np

# read a given column(col_idx) from `row_min`(inclusive) to `row_max`(inclusive) to numpy array
# --> `row_min` and `row_max` are based on the row index in Excel
def readCoulmn( sheet, col_idx, row_min, row_max ):
    array1D = np.array([])

    for i in range(row_min, row_max + 1):
        cell = sheet.cell(row = i, column = col_idx)
        array1D = np.append(array1D, cell.value)

    return array1D


############################################
### Reading
############################################
ExcelFile = "uniform-distribution.xlsx"
wb = pyxl.load_workbook(ExcelFile, data_only=True)
sheet = wb.active

col_max  = sheet.max_column
row_max  = sheet.max_row

UniDist1 = np.array([])
UniDist2 = np.array([])

UniDist1 = readCoulmn(sheet, 1, 2, row_max)
UniDist2 = readCoulmn(sheet, 2, 2, row_max)


############################################
### Transformation
############################################
# We use Box-Muller transformation to obtain two Gaussians from two uniform distribution
# --> https://en.wikipedia.org/wiki/Normal_distribution#Generating_values_from_normal_distribution
# --> https://en.wikipedia.org/wiki/Box%E2%80%93Muller_transform
NorDist1 = np.sqrt(-2*np.log(UniDist1))*np.cos(2*np.pi*UniDist2)
NorDist2 = np.sqrt(-2*np.log(UniDist1))*np.sin(2*np.pi*UniDist2)


############################################
### Plotting
############################################
x = np.arange(2, row_max+1, 1)

plt.scatter(x, NorDist1, s=20, c="g", marker=r'$\clubsuit$', label="NorDist1")
plt.scatter(x, NorDist2, s=20, c="r", marker=r'$\clubsuit$', label="NorDist2")

plt.xlabel("Row index")
plt.ylabel("Something")
plt.legend(loc='upper left')
plt.savefig( "fig__Gaussian.png", bbox_inches='tight', pad_inches=0, format="png" )
plt.show()
